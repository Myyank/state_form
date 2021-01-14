from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter,cell 
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging

master = Tk()
master.title("Form Creator")
master.minsize(640,400)


#backend code starts here

systemdrive = os.getenv('WINDIR')[0:3]
dbfolder = os.path.join(systemdrive,'Forms\DB')
#dbfolder = "D:\Company Projects\Form creator\DB"
State_forms = os.path.join(systemdrive,'Forms\State forms')
#State_forms = "D:\Company Projects\Form creator\State forms"
Statefolder = Path(State_forms)
logfolder = os.path.join(systemdrive,'Forms\logs')
#logfolder = "D:\Company Projects\Form creator\logs"


monthdict= {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}


log_filename = datetime.datetime.now().strftime(os.path.join(logfolder,'logfile_%d_%m_%Y_%H_%M_%S.log'))
logging.basicConfig(filename=log_filename, level=logging.INFO)

def create_pdf(folderlocation,file_name):
    import win32com.client
    from pywintypes import com_error



    excel_filename = file_name
    pdf_filename = file_name.split('.')[0]+'.pdf'

    


    # Path to original excel file
    WB_PATH=os.path.join(folderlocation,excel_filename)
    # PDF path when saving
    PATH_TO_PDF =os.path.join(folderlocation,pdf_filename)

    logging.info(WB_PATH)
    logging.info(PATH_TO_PDF)


    excel = win32com.client.Dispatch("Excel.Application")

    excel.Visible = False

    try:
        logging.info('Start conversion to PDF')

        # Open
        wb = excel.Workbooks.Open(WB_PATH)

        sheetnumbers= len(pd.ExcelFile(WB_PATH).sheet_names)

        # Specify the sheet you want to save by index. 1 is the first (leftmost) sheet.
        ws_index_list = list(range(1,sheetnumbers+1))
        wb.WorkSheets(ws_index_list).Select()

        # Save
        wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
    except com_error as e:
        logging.info('failed.')
    else:
        logging.info('Succeeded.')
    finally:
        wb.Close()
        excel.Quit()

def Maharashtra(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info('Maharashtra forms')

def Delhi(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info('Delhi forms')

def Tamilnadu(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info('Tamilnadu forms')

def Telangana(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info('Telangana forms')

def Uttar_Pradesh(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info('Uttar Pradesh forms')
    UPfilespath = os.path.join(Statefolder,'Uttar Pradesh')
    logging.info('Uttar Pradesh files path is :'+str(UPfilespath))
    data.reset_index(drop=True, inplace=True)

    month_num = monthdict[month]

    def create_form_D():

        formDfilepath = os.path.join(UPfilespath,'Form D.xlsx')
        formDfile = load_workbook(filename=formDfilepath)
        logging.info('Form D file has sheet: '+str(formDfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formD = data.copy()

        data_formD['Serial No.'] = list(range(1,len(data_formD)+1))
        data_formD['Sign']="TestSign"
        data_formD['Remarks']="TestRemark"
        colToAdd = ['Wage Rate','Deductions_amt','Reason_Deduction','Date_Deduction','Deduction_dt','Deductions_np','No_Instalments','Amount_Realized_dt','Amount_Realized_amt','Amount_Realized_np']
        data_formD = data_formD.reindex(data_formD.columns.tolist() + colToAdd, axis=1)
        data_formD.loc[:,colToAdd] = '----'
        formD_columns = ["Serial No.",'Employee Name','Wage Rate','Deduction_dt','Deductions_amt','Deductions_np','Reason_Deduction','Date_Deduction','No_Instalments','Amount_Realized_dt','Amount_Realized_amt','Amount_Realized_np','Remarks','Sign']
        formD_data = data_formD[formD_columns]
        

        formDsheet = formDfile['Sheet1']

        formDsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form D is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formD_data, index=False, header=False)

        logging.info('rows taken out from data')
    
        

        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                formDsheet.cell(row=r_idx, column=c_idx, value=value)
                formDsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Calibri', size =10)
                formDsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formDsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
            formDsheet.insert_rows(r_idx+1,1)    
        logging.info('')
        
            
        
        '''establishment = formAsheet['L6'].value
        L6_data = establishment+' '+data_formA['Unit'][0]+', '+data_formA['Branch'][0]
        formAsheet['L6'] = L6_data

        company = formAsheet['A10'].value
        A10_data = company+' '+data_formA['Unit'][0]+', '+data_formA['Branch'][0]
        formAsheet['A10'] = A10_data'''
        monthstart = datetime.date(year,month_num,1)
        monthend = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
        formDsheet['A4'].value=None
        formDsheet.unmerge_cells('A4:N4')
        formDsheet['A4'].value='(From ' + str(monthstart)+ " To " + str(monthend)+ ' )'
        formDsheet.merge_cells('A4:N4')   
        for row_id in range(1,4):
                for col_id in range(1,15):
                    formDsheet.cell(row=row_id, column=col_id).border = Border(left=Side(style='medium'), 
                                                                                       right=Side(style='medium'), 
                                                                                           top=Side(style='medium'), 
                                                                                         bottom=Side(style='medium'))
        for row_id in range(6,r_idx+1):
                for col_id in range(1,15):
                    formDsheet.cell(row=row_id, column=col_id).border = Border(left=Side(style='medium'), 
                                                                                       right=Side(style='medium'), 
                                                                                           top=Side(style='medium'), 
                                                                                         bottom=Side(style='medium'))
           
        formDfinalfile = os.path.join(filelocation,'Form D.xlsx')
        formDfile.save(filename=formDfinalfile)
        
    def create_form_G():
        formGfilepath = os.path.join(UPfilespath,'Form G.xlsx')
        formGfile = load_workbook(filename=formGfilepath)
        logging.info('Form G file has sheet: '+str(formGfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formG = data.copy()
        colToAdd = ['Date','Rest_Start','Rest_Ends','D_F_A','blank','Location Code','OT hours' 'OVERTIME' 'Sign' 'Amount' 'PAY_DATE' 'Sign']
        data_formG = data_formG.reindex(data_formG.columns.tolist() + colToAdd, axis=1)
        data_formG['Date'] = month
        data_formG['Rest_Start'] = 'Rest_Start'
        data_formG['Rest_Ends'] = 'Rest_Ends'
        data_formG['D_F_A'] = 'D_F_A'
        data_formG['blank'] = 'blank'
        data_formG['Location Code'] = 'Location Code'
        data_formG['OT hours'] = 'OT hours'
        data_formG['OVERTIME'] = 'OVERTIME'
        data_formG['Sign'] = 'Sign'
        data_formG['Amount'] = 'Amount'
        data_formG['PAY_DATE'] = 'PAY_DATE'
        
        formG_columns = ['Date','start_time','Rest_Start','Rest_Ends','end_time','OT hours','Earned Basic','D_F_A','OVERTIME','Sign','Amount','PAY_DATE','Salary Advance','blank','Other Deduction','Net Paid','Sign','Employee Name']

        formG_data = data_formG[formG_columns]

        formGsheet = formGfile['Sheet1']

        formGsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form G is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formG_data, index=False, header=False)
        temp = dataframe_to_rows(formG_data, index=False, header=True)
        logging.info('rows taken out from data')
        #print (rows)
        '''for r_idx, row in enumerate(rows, 12):
            #print(r_idx," ",row)
            empName = data_formG.loc[temp, "Employee Name"]
            new = formGfile.copy_worksheet(formGsheet)
            new.title = empName
            idx=idx+1
            for c_idx, value in enumerate(row, 1):
                
                formGsheet.cell(row=r_idx, column=c_idx, value=value)
                formGsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Calibri', size =10)
                border_sides = Side(style='thin')
                formGsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            #formGsheet.insert_rows(r_idx+1,1)
            '''
        form_data = data.copy()
            
        for idx in form_data.index:
            empName = form_data.loc[idx, "Employee Name"]
            

                # create a new sheet for employee
            new = formGfile.copy_worksheet(formGsheet)
            new.title = empName
            c_idx=1
            new.cell(row= 12, column=c_idx, value=month)
            new.cell(row= 12, column=c_idx+1, value=form_data.loc[idx, "start_time"])  
            new.cell(row= 12, column=c_idx+2, value="Rest_Start") 
            new.cell(row= 12, column=c_idx+3, value="Rest_Ends") 
            new.cell(row= 12, column=c_idx+4, value=form_data.loc[idx, "end_time"]) 
            new.cell(row= 12, column=c_idx+5, value="OT hours") # earned basic B11
            new.cell(row= 12, column=c_idx+6, value=form_data.loc[idx, "Earned Basic"])# House rent allowance B13
            new.cell(row= 12, column=c_idx+7, value="D_F_A") # overtime wages B14
            new.cell(row= 12, column=c_idx+8, value="OVERTIME") # otehr allowance B16
            new.cell(row= 12, column=c_idx+9, value="Sign") # Gross wages B17
            new.cell(row= 12, column=c_idx+10, value="Amount") # employee state insurance I12
            new.cell(row= 12, column=c_idx+11, value="PAY_DATE") # other deductions I13
            new.cell(row= 12, column=c_idx+13, value=form_data.loc[idx, "Salary Advance"]) # net paid #G17
            new.cell(row= 12, column=c_idx+14, value="blank")
            new.cell(row= 12, column=c_idx+15, value=form_data.loc[idx, "Other Deduction"])
            new.cell(row= 12, column=c_idx+16, value=form_data.loc[idx, "Net Paid"])
            new.cell(row= 12, column=c_idx+17, value="Sign")
            new.cell(row=4, column=1,value="Name of employee - "+ form_data.loc[idx, "Employee Name"]+". Man/Woman/Young Person/Child")
            new.cell(row=5,column=1,value="Father’s/Husband’s name -" +form_data.loc[idx,"Father's Name"]+". Address - "+ form_data.loc[idx,"Address"] +"Nature of employment ………………. Whether employed on daily, monthly, contract or piece-rate wages with rate………………………….. Wage period………………   date of Employment………………")
            for row_id in range(1,6):
                for col_id in range(1,19):
                    new.cell(row=row_id, column=col_id).border = Border(left=Side(style='thin'), 
                                                                                       right=Side(style='thin'), 
                                                                                           top=Side(style='thin'), 
                                                                                         bottom=Side(style='thin'))
            for row_id in range(7,15):
                for col_id in range(1,19):
                    new.cell(row=row_id, column=col_id).border = Border(left=Side(style='medium'), 
                                                                                       right=Side(style='medium'), 
                                                                                           top=Side(style='medium'), 
                                                                                         bottom=Side(style='medium'))
            
        
        formGfile.remove(formGfile['Sheet1']) 
        formGfinalfile = os.path.join(filelocation,'Form G.xlsx')
        formGfile.save(filename=formGfinalfile)
    
   
    def create_form_H():
        formHfilepath = os.path.join(UPfilespath,'Form H.xlsx')
        formHfile = load_workbook(filename=formHfilepath)
        logging.info('Form H file has sheet: '+str(formHfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formH = data.copy()

        formHsheet = formHfile['Sheet1']
        formHsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form H is ready')

        for idx in data_formH.index:

            empName = data_formH.loc[idx, "Employee Name"]
            #print("Employee:", empName)

            # create a new sheet for employee
            new = formHfile.copy_worksheet(formHsheet)
            new.title = empName
            
            monthstart = datetime.date(year,month_num,1)
            monthend = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
            
            new.cell(row= 4, column=1, value='Period from ' + str(monthstart)+ " To " + str(monthend))
            new.merge_cells('A4:D4') 
            # popoluate every cell with required employee information 
            new.cell(row= 10, column=1, value=data_formH.loc[idx, "Opening"])    
            for row_id in range(1,4):
                for col_id in range(1,15):
                    new.cell(row=row_id, column=col_id).border = Border(left=Side(style='medium'), 
                                                                                       right=Side(style='medium'), 
                                                                                           top=Side(style='medium'), 
                                                                                         bottom=Side(style='medium'))
            for row_id in range(6,14):
                for col_id in range(1,15):
                    new.cell(row=row_id, column=col_id).border = Border(left=Side(style='medium'), 
                                                                                       right=Side(style='medium'), 
                                                                                           top=Side(style='medium'), 
                                                                                         bottom=Side(style='medium'))
                    
        
        formHfile.remove(formHfile['Sheet1'])
        formHfinalfile = os.path.join(filelocation,'Form H.xlsx')
        formHfile.save(filename=formHfinalfile)

    def create_form_XIV():
        formXIVfilepath = os.path.join(UPfilespath,'Form XIV.xlsx')
        formXIVfile = load_workbook(filename=formXIVfilepath)
        logging.info('Form XX file has sheet: '+str(formXIVfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXIV = data.copy()

        ''''data_formXX['S.no'] = list(range(1,len(data_formXX)+1))

        data_formXX['a'] ='---'
        data_formXX['b'] ='---'
        data_formXX['c'] ='---'
        data_formXX['d'] ='---'
        data_formXX['e'] ='---'
        data_formXX['f'] ='---'
        data_formXX['g'] ='---'
        data_formXX['h'] ='---'
        data_formXX['i'] =''

        formXX_columns = ['S.no','Employee Name',"Father's Name",'Designation','a','b','c','d','e','f','g','h','i']

        formXX_data = data_formXX[formXX_columns]

        formXXsheet = formXXfile['FORM XX'] '''
        
        formXIVsheet = formXIVfile['Sheet1']
        formXIVsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XIV is ready')
        
        data_formXIV['Serial No.'] = list(range(1,len(data_formXIV)+1))
        sr=1
        for idx in data_formXIV.index:

        
            formXIVsheet.cell(row= 5, column=1, value="1.Name of the employment:-"+data_formXIV.loc[idx,"Unit"])
            formXIVsheet.cell(row= 6, column=1, value="2.Name of establishment or concern :-"+data_formXIV.loc[idx,"Unit"])
            formXIVsheet.cell(row= 7, column=1, value="3. Address:-"+data_formXIV.loc[idx,"Address"])
            # popoluate every cell with required employee information 
            if pd.isnull(data_formXIV.loc[idx,"Emp Name"]):
                continue
            else:
                formXIVsheet.cell(row=9+sr, column=1).font =Font(name ='Calibri', size =10,bold=False)
                formXIVsheet.cell(row=9+sr, column=1).alignment = Alignment(horizontal='justify', vertical='justify', wrap_text = True)
                formXIVsheet.cell(row=9+sr, column=2).font =Font(name ='Calibri', size =10,bold=False)
                formXIVsheet.cell(row=9+sr, column=3).font =Font(name ='Calibri', size =10,bold=False)
                formXIVsheet.insert_rows(9+sr+2,1) 
                formXIVsheet.cell(row=9+sr+2, column=3).border = Border(right=Side(style='medium')) 
                formXIVsheet.cell(row= 9+sr, column=1, value=sr)
                formXIVsheet.cell(row= 9+sr, column=2, value=data_formXIV.loc[idx,"Emp Name"])
                formXIVsheet.cell(row= 9+sr, column=3, value="Mapping Not Found")
                for col_id in range(1,4):
                    formXIVsheet.cell(row=9+sr, column=col_id).border = Border(left=Side(style='medium'), 
                                                                                        right=Side(style='medium'), 
                                                                                            top=Side(style='medium'), 
                                                                                            bottom=Side(style='medium'))
                sr=sr+1
        if sr>2:
            formXIVsheet.delete_rows(9+sr+1,2)
        else :
            formXIVsheet.delete_rows(9+sr+1,1)
        
        for row_id in range(1,5):
                for col_id in range(1,4):
                    formXIVsheet.cell(row=row_id, column=col_id).border = Border(left=Side(style='medium'), 
                                                                                       right=Side(style='medium'), 
                                                                                           top=Side(style='medium'), 
                                                                                         bottom=Side(style='medium'))
        logging.info('rows taken out from data')
               
        formXIVfinalfile = os.path.join(filelocation,'Form XIV.xlsx')
        formXIVfile.save(filename=formXIVfinalfile)

    def create_formXVII():
        formXVIIfilepath = os.path.join(UPfilespath,'Form XVII.xlsx')
        formXVIIfile = load_workbook(filename=formXVIIfilepath)
        logging.info('wages file has sheet: '+str(formXVIIfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXVII = data.copy()

        
        formXVIIsheet = formXVIIfile['Sheet1']
        formXVIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        

        logging.info('data for wages is ready')
        sr=1
        formXVIIsheet.cell(row= 4, column=1, value='Month ending '+str(month)+'. '+str(year))
        for idx in data_formXVII.index:

            if pd.isnull(data_formXVII.loc[idx,"Emp Name"]):
                continue
            else:
                for col in range(1,17):
                    formXVIIsheet.cell(row=3+sr*4, column=col).font =Font(name ='Calibri', size =10,bold=False)
                    formXVIIsheet.cell(row=3+sr*4, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                formXVIIsheet.cell(row= 3+sr*4, column=1, value=sr)
                formXVIIsheet.cell(row= 3+sr*4, column=2, value=data_formXVII.loc[idx,"Emp Name"])
                formXVIIsheet.cell(row= 3+sr*4, column=3, value=data_formXVII.loc[idx,"Father's Name"])
                formXVIIsheet.cell(row= 3+sr*4, column=5, value=data_formXVII.loc[idx,"Gender"])
                formXVIIsheet.cell(row= 3+sr*4, column=6, value=data_formXVII.loc[idx,"DesigName"])
                formXVIIsheet.cell(row= 3+sr*4, column=7, value="Mapping not found")
                formXVIIsheet.cell(row= 3+sr*4, column=8, value="---")
                formXVIIsheet.cell(row= 3+sr*4, column=9, value="---")
                formXVIIsheet.cell(row= 3+sr*4, column=10, value="8 hrs")
                formXVIIsheet.cell(row= 3+sr*4, column=11, value=data_formXVII.loc[idx,"FIXED MONTHLY GROSS"])
                formXVIIsheet.cell(row= 3+sr*4, column=12, value="Double")
                formXVIIsheet.cell(row= 3+sr*4, column=13, value="Mapping not found")
                formXVIIsheet.cell(row= 3+sr*4, column=14, value="OVERTIME")
                formXVIIsheet.cell(row= 3+sr*4, column=15, value="GROSS_EARN")
                formXVIIsheet.cell(row= 3+sr*4, column=16, value=data_formXVII.loc[idx,"CHECK CTC Gross"])
                sr=sr+1
                for c_idx in range(1,17):
                    if int(c_idx)!=3 and int(c_idx)!=4:
                        formXVIIsheet.merge_cells(str(get_column_letter(c_idx))+str(3+sr*4)+':'+str(get_column_letter(c_idx))+str(6+sr*4))
                    elif int(c_idx)==3:
                        formXVIIsheet.merge_cells('C'+str(3+sr*4)+':D'+str(6+sr*4))
                    else:
                        continue

        
        logging.info('rows taken out from data')
        for row_id in range(5,sr*4+3):
                for col_id in range(1,17):
                    formXVIIsheet.cell(row=row_id, column=col_id).border = Border(left=Side(style='medium'), 
                                                                                       right=Side(style='medium'), 
                                                                                           top=Side(style='medium'), 
                                                                                         bottom=Side(style='medium'))
        for row_id in range(1,5):
                for col_id in range(1,17):
                    formXVIIsheet.cell(row=row_id, column=col_id).border = Border(left=Side(style='thin'), 
                                                                                       right=Side(style='thin'), 
                                                                                           top=Side(style='thin'), 
                                                                                         bottom=Side(style='thin'))
        
                              
        formXVIIfinalfile = os.path.join(filelocation,'Form XVII.xlsx')
        formXVIIfile.save(filename=formXVIIfinalfile)

    
    def create_form_C():
        formCfilepath = os.path.join(UPfilespath,'Form C.xlsx')
        formCfile = load_workbook(filename=formCfilepath)
        logging.info('wages file has sheet: '+str(formCfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formC = data.copy()

        formCsheet = formCfile['Sheet1']
        formCsheet.sheet_properties.pageSetUpPr.fitToPage = True
        for idx in data_formC.index:
            if pd.isnull(data_formC.loc[idx,"Employee Name"]):
                continue
            else :
                empName = data_formC.loc[idx, "Employee Name"]
                #print("Employee:", empName)

                # create a new sheet for employee
                new = formCfile.copy_worksheet(formCsheet)
                new.title = empName
                for row_id in range(1,4):
                    for col_id in range(1,16):
                        new.cell(row=row_id, column=col_id).border = Border(left=Side(style='medium'), 
                                                                                        right=Side(style='medium'), 
                                                                                            top=Side(style='medium'), 
                                                                                            bottom=Side(style='medium'))
                for row_id in range(4,9):
                        for col_id in range(1,16):
                            new.cell(row=row_id, column=col_id).border = Border(left=Side(style=None), 
                                                                                            right=Side(style=None), 
                                                                                                top=Side(style=None), 
                                                                                                bottom=Side(style=None))
                for row_id in range(9,13):
                        for col_id in range(1,16):
                            new.cell(row=row_id, column=col_id).border = Border(left=Side(style='medium'), 
                                                                                            right=Side(style='medium'), 
                                                                                                top=Side(style='medium'), 
                                                                                                bottom=Side(style='medium'))
                
                new.cell(row=5 , column=1, value="Name of employee:- "+str(data_formC.loc[idx,"Employee Name"])+". Man/Woman/young person/child")
                new.cell(row=6 , column=1, value="Father’s/Husband’s Name:- "+str(data_formC.loc[idx,"Father's Name"])+". Address:-"+str(data_formC.loc[idx,"Address"]))
                new.cell(row=7 , column=1, value="Nature of employment:- "+str(data_formC.loc[idx,"Designation"]))
                new.cell(row=8 , column=1, value="Whether employed on daily, monthly, contract or piece-rate wages with rate  "+str(month))
                
                for col in range(1,16):
                    new.cell(row=12, column=col).font =Font(name ='Calibri', size =10,bold=False)
                    new.cell(row=12, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                new.cell(row=12 , column=1, value=str(month))
                new.cell(row=12 , column=3, value=data_formC.loc[idx,"Opening"])
                new.cell(row=12 , column=4, value=data_formC.loc[idx,"Monthly Increment"])
                new.cell(row=12 , column=5, value="No Mapping Found")
                new.cell(row=12 , column=6, value=data_formC.loc[idx,"Leave Accrued"])
                new.cell(row=12 , column=7, value="---")
                new.cell(row=12 , column=8, value="---")
                new.cell(row=12 , column=9, value="OT Hours")
                new.cell(row=12 , column=10, value=data_formC.loc[idx,"FIXED MONTHLY GROSS"])
                new.cell(row=12 , column=11, value=data_formC.loc[idx,"Salary Advance"])
                new.cell(row=12 , column=12, value="No Mapping Found")
                new.cell(row=12 , column=13, value=data_formC.loc[idx,"Other Deduction"])              
                new.cell(row=12 , column=14, value=data_formC.loc[idx,"Total Deductions"]) 
                new.cell(row=12 , column=15, value=data_formC.loc[idx,"Net Paid"]) 
                for row_id in range (4,9):
                    new.cell(row=row_id, column=15).border = Border(left=Side(style=None), 
                                                                                            right=Side(style='medium'), 
                                                                                                top=Side(style=None), 
                                                                                                bottom=Side(style=None))
            
        formCfile.remove(formCfile['Sheet1'])
        formCfinalfile = os.path.join(filelocation,'Form C.xlsx')
        formCfile.save(filename=formCfinalfile)
        
        
        
    create_form_D()
    create_form_G()
    create_form_H()
    create_form_XIV() 
    create_formXVII()
    create_form_C()
    '''create_form_B()
    create_form_XXI()
    create_form_XXII()
    create_form_XXIII()
    create_form_XX()
    create_wages()
    create_form_H_F('FORM H')
    create_form_H_F('FORM F')
    create_muster()
    create_formXIX()
    create_ecard()'''


def Karnataka(data,contractor_name,contractor_address,filelocation,month,year):
    karnatakafilespath = os.path.join(Statefolder,'Karnataka')
    logging.info('karnataka files path is :'+str(karnatakafilespath))
    data.reset_index(drop=True, inplace=True)

    month_num = monthdict[month]

    def create_form_A():

        formAfilepath = os.path.join(karnatakafilespath,'FormA.xlsx')
        formAfile = load_workbook(filename=formAfilepath)
        logging.info('Form A file has sheet: '+str(formAfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formA = data.copy()

        data_formA['S.no'] = list(range(1,len(data_formA)+1))

        data_formA["Nationality"] =''
        data_formA['education level'] = ''
        data_formA['Category address'] = ''
        data_formA['type of employment'] = ''
        data_formA['lwf'] = ''
        data_formA['Service Book No'] = ''
        formA_columns = ["S.no",'Employee Code','Employee Name','Unit','Location',"Gender","Father's Name",'Date of Birth',"Nationality","education level",'Date Joined','Designation',"Category address","type of employment",'Mobile Tel No.','UAN Number',"PAN Number","ESIC Number","lwf","Aadhar Number","Bank A/c Number","Bank Name","Account Code","P","L","Service Book No","Date Left","Reason for Leaving"]
        formA_data = data_formA[formA_columns]


        formAsheet = formAfile['FORM A']

        formAsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form A is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formA_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 13):
            for c_idx, value in enumerate(row, 1):
                formAsheet.cell(row=r_idx, column=c_idx, value=value)
                formAsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formAsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formAsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        logging.info('')

        establishment = formAsheet['L6'].value
        L6_data = establishment+' '+data_formA['Unit'][0]+', '+data_formA['Branch'][0]
        formAsheet['L6'] = L6_data

        company = formAsheet['A10'].value
        A10_data = company+' '+data_formA['Unit'][0]+', '+data_formA['Branch'][0]
        formAsheet['A10'] = A10_data

        formAfinalfile = os.path.join(filelocation,'FormA.xlsx')
        formAfile.save(filename=formAfinalfile)
        

    def create_form_B():
        formBfilepath = os.path.join(karnatakafilespath,'FormB.xlsx')
        formBfile = load_workbook(filename=formBfilepath)
        logging.info('Form B file has sheet: '+str(formBfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formB = data.copy()

        data_formB['OT hours'] = 0
        data_formB['Pay OT'] = 0
        data_formB['basic_and_allo'] = data_formB['Earned Basic']+ data_formB['Other Allowance']+ data_formB['Meal Allowance'] +data_formB['Special Allowance'] +data_formB['Personal Allowance']
        data_formB['Other EAR'] = data_formB['Other Reimb']+data_formB['Arrears']+data_formB['Other Earning']+data_formB['Variable Pay']+data_formB['Stipend'] +data_formB['Consultancy Fees']
        data_formB['VPF']=0
        data_formB['Income Tax']=0
        data_formB['EMP PF'] = data_formB['PF']
        data_formB['BankID'] = ''
        data_formB['Pay Date'] = ''
        data_formB['Remarks'] =''

        formB_columns = ['Employee Code','Employee Name','FIXED MONTHLY GROSS',	'Days Paid','OT hours',	'basic_and_allo', 'Pay OT',	'HRA',	'Tel and Int Reimb', 'Bonus', 'Fuel Reimb',	'Prof Dev Reimb', 'Corp Attire Reimb',	'CCA',	'Leave Encashment',	'Other EAR', 'Total Earning', 'PF',	'ESIC',	'VPF', 'Loan Deduction', 'Loan Interest', 'P.Tax',	'CSR',	'Income Tax', 'Insurance',	'LWF EE',	'Other Deduction',	'TDS',	'Total Deductions',	'Net Paid',	'EMP PF','BankID','Pay Date','Remarks']

        formB_data = data_formB[formB_columns]

        formBsheet = formBfile['FORM B']

        formBsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form B is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formB_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 21):
            for c_idx, value in enumerate(row, 2):
                formBsheet.cell(row=r_idx, column=c_idx, value=value)
                formBsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                border_sides = Side(style='thin')
                formBsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formBsheet['B10'].value
        B10_data = contractline+' '+contractor_name+', '+contractor_address
        formBsheet['B10'] = B10_data

        locationline = formBsheet['B11'].value
        B11_data = locationline+' '+data_formB['Unit'][0]+', '+data_formB['Branch'][0]
        formBsheet['B11'] = B11_data

        establine = formBsheet['B12'].value
        B12_data = establine+' '+data_formB['Unit'][0]+', '+data_formB['Branch'][0]
        formBsheet['B12'] = B12_data

        peline = formBsheet['B13'].value
        B13_data = peline+' '+data_formB['Unit'][0]+', '+data_formB['Branch'][0]
        formBsheet['B13'] = B13_data

        monthstart = datetime.date(year,month_num,1)
        monthend = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
        formBsheet['B16'] = 'Wage period From: '+str(monthstart)+' to '+str(monthend)

        formBfinalfile = os.path.join(filelocation,'FormB.xlsx')
        formBfile.save(filename=formBfinalfile)

    def create_form_XXI():
        formXXIfilepath = os.path.join(karnatakafilespath,'FormXXI.xlsx')
        formXXIfile = load_workbook(filename=formXXIfilepath)
        logging.info('Form XXI file has sheet: '+str(formXXIfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXXI = data.copy()

        data_formXXI['S.no'] = list(range(1,len(data_formXXI)+1))

        data_formXXI['a'] ='---'
        data_formXXI['b'] ='---'
        data_formXXI['c'] ='---'
        data_formXXI['e'] ='---'
        data_formXXI['f'] ='---'
        data_formXXI['g'] =''

        formXXI_columns = ['S.no','Employee Name',"Father's Name",'Designation','a','b','c','FIXED MONTHLY GROSS','e','f','g']

        formXXI_data = data_formXXI[formXXI_columns]

        formXXIsheet = formXXIfile['FORM XXI']

        formXXIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXI is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXXI_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 14):
            for c_idx, value in enumerate(row, 3):
                formXXIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                border_sides = Side(style='thin')
                formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIsheet['C7'].value
        C7_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIsheet['C7'] = C7_data

        locationline = formXXIsheet['C8'].value
        C8_data = locationline+' '+data_formXXI['Unit'][0]+', '+data_formXXI['Branch'][0]
        formXXIsheet['C8'] = C8_data

        establine = formXXIsheet['C9'].value
        C9_data = establine+' '+data_formXXI['Unit'][0]+', '+data_formXXI['Branch'][0]
        formXXIsheet['C9'] = C9_data

        peline = formXXIsheet['C10'].value
        C10_data = peline+' '+data_formXXI['Unit'][0]+', '+data_formXXI['Branch'][0]
        formXXIsheet['C10'] = C10_data

        formXXIfinalfile = os.path.join(filelocation,'FormXXI.xlsx')
        formXXIfile.save(filename=formXXIfinalfile)



    def create_form_XXII():
        formXXIIfilepath = os.path.join(karnatakafilespath,'FormXXII.xlsx')
        formXXIIfile = load_workbook(filename=formXXIIfilepath)
        logging.info('Form XXII file has sheet: '+str(formXXIIfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXXII = data.copy()

        data_formXXII['S.no'] = list(range(1,len(data_formXXII)+1))

        data_formXXII['b'] ='---'
        data_formXXII['c'] ='---'
        data_formXXII['d'] ='---'
        data_formXXII['e'] ='---'
        data_formXXII['f'] ='---'
        data_formXXII['g'] =''

        formXXII_columns = ['S.no','Employee Name',"Father's Name",'Designation','FIXED MONTHLY GROSS','b','c','d','e','f','g']

        formXXII_data = data_formXXII[formXXII_columns]

        formXXIIsheet = formXXIIfile['FORM XXII']

        formXXIIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXII is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXXII_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 15):
            for c_idx, value in enumerate(row, 3):
                formXXIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                border_sides = Side(style='thin')
                formXXIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIIsheet['C7'].value
        C7_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIIsheet['C7'] = C7_data

        locationline = formXXIIsheet['C8'].value
        C8_data = locationline+' '+data_formXXII['Unit'][0]+', '+data_formXXII['Branch'][0]
        formXXIIsheet['C8'] = C8_data

        establine = formXXIIsheet['C9'].value
        C9_data = establine+' '+data_formXXII['Unit'][0]+', '+data_formXXII['Branch'][0]
        formXXIIsheet['C9'] = C9_data

        peline = formXXIIsheet['C10'].value
        C10_data = peline+' '+data_formXXII['Unit'][0]+', '+data_formXXII['Branch'][0]
        formXXIIsheet['C10'] = C10_data

        formXXIIfinalfile = os.path.join(filelocation,'FormXXII.xlsx')
        formXXIIfile.save(filename=formXXIIfinalfile)


    def create_form_XXIII():
        formXXIIIfilepath = os.path.join(karnatakafilespath,'FormXXIII.xlsx')
        formXXIIIfile = load_workbook(filename=formXXIIIfilepath)
        logging.info('Form XXIII file has sheet: '+str(formXXIIIfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXXIII = data.copy()

        data_formXXIII['S.no'] = list(range(1,len(data_formXXIII)+1))

        data_formXXIII['b'] ='---'
        data_formXXIII['c'] ='---'
        data_formXXIII['d'] ='---'
        data_formXXIII['e'] ='---'
        data_formXXIII['f'] ='---'
        data_formXXIII['g'] =''

        formXXIII_columns = ['S.no','Employee Name',"Father's Name",'Designation','FIXED MONTHLY GROSS','b','c','d','e','f','g']

        formXXIII_data = data_formXXIII[formXXIII_columns]

        formXXIIIsheet = formXXIIIfile['FORM XXIII']

        formXXIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXIII is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXXIII_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 12):
            for c_idx, value in enumerate(row, 3):
                formXXIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                border_sides = Side(style='thin')
                formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIIIsheet['C5'].value
        C5_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIIIsheet['C5'] = C5_data

        locationline = formXXIIIsheet['C6'].value
        C6_data = locationline+' '+data_formXXIII['Unit'][0]+', '+data_formXXIII['Branch'][0]
        formXXIIIsheet['C6'] = C6_data

        establine = formXXIIIsheet['C7'].value
        C7_data = establine+' '+data_formXXIII['Unit'][0]+', '+data_formXXIII['Branch'][0]
        formXXIIIsheet['C7'] = C7_data

        peline = formXXIIIsheet['C8'].value
        C8_data = peline+' '+data_formXXIII['Unit'][0]+', '+data_formXXIII['Branch'][0]
        formXXIIIsheet['C8'] = C8_data

        formXXIIIfinalfile = os.path.join(filelocation,'FormXXIII.xlsx')
        formXXIIIfile.save(filename=formXXIIIfinalfile)


    def create_form_XX():
        formXXfilepath = os.path.join(karnatakafilespath,'FormXX.xlsx')
        formXXfile = load_workbook(filename=formXXfilepath)
        logging.info('Form XX file has sheet: '+str(formXXfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXX = data.copy()

        data_formXX['S.no'] = list(range(1,len(data_formXX)+1))

        data_formXX['a'] ='---'
        data_formXX['b'] ='---'
        data_formXX['c'] ='---'
        data_formXX['d'] ='---'
        data_formXX['e'] ='---'
        data_formXX['f'] ='---'
        data_formXX['g'] ='---'
        data_formXX['h'] ='---'
        data_formXX['i'] =''

        formXX_columns = ['S.no','Employee Name',"Father's Name",'Designation','a','b','c','d','e','f','g','h','i']

        formXX_data = data_formXX[formXX_columns]

        formXXsheet = formXXfile['FORM XX']

        formXXsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XX is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXX_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 15):
            for c_idx, value in enumerate(row, 3):
                formXXsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                border_sides = Side(style='thin')
                formXXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXsheet['C6'].value
        C6_data = contractline+' '+contractor_name+', '+contractor_address
        formXXsheet['C6'] = C6_data

        locationline = formXXsheet['C7'].value
        C7_data = locationline+' '+data_formXX['Unit'][0]+', '+data_formXX['Branch'][0]
        formXXsheet['C7'] = C7_data

        establine = formXXsheet['C8'].value
        C8_data = establine+' '+data_formXX['Unit'][0]+', '+data_formXX['Branch'][0]
        formXXsheet['C8'] = C8_data

        peline = formXXsheet['C9'].value
        C9_data = peline+' '+data_formXX['Unit'][0]+', '+data_formXX['Branch'][0]
        formXXsheet['C9'] = C9_data

        formXXfinalfile = os.path.join(filelocation,'FormXX.xlsx')
        formXXfile.save(filename=formXXfinalfile)

    def create_wages():
        wagesfilepath = os.path.join(karnatakafilespath,'Wages.xlsx')
        wagesfile = load_workbook(filename=wagesfilepath)
        logging.info('wages file has sheet: '+str(wagesfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_wages = data.copy()

        data_wages['S.no'] = list(range(1,len(data_wages)+1))

        data_wages['Site Address'] = ''

        data_wages['OT hours'] = 0
        data_wages['Conv'] = 0
        data_wages['Sub Allow'] = 0
        data_wages['Fines']=0
        data_wages['Damages']=0
        data_wages['Pay mode'] = ''
        data_wages['Remarks'] =''

        wages_columns = ['S.no','Employee Code','Employee Name','Gender','Designation','Department','Site Address','Date Joined','ESIC Number','PF Number','FIXED MONTHLY GROSS','Days Paid','OT hours','Earned Basic','HRA','Conv','Medical Allowance','Tel and Int Reimb', 'Bonus', 'Fuel Reimb',	'Prof Dev Reimb', 'Corp Attire Reimb','Special Allowance',	'CCA','Other Earning',	'Sub Allow','Leave Encashment', 'Total Earning','ESIC', 'PF','P.Tax','TDS','CSR','Insurance','Salary Advance','Fines','Damages','Other Deduction',	'Total Deductions',	'Net Paid','Pay mode','Bank A/c Number','Remarks']

        wages_data = data_wages[wages_columns]

        wagessheet = wagesfile['Wages']

        wagessheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for wages is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(wages_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 21):
            for c_idx, value in enumerate(row, 2):
                wagessheet.cell(row=r_idx, column=c_idx, value=value)
                wagessheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                border_sides = Side(style='thin')
                wagessheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = wagessheet['A10'].value
        A10_data = contractline+' '+contractor_name+', '+contractor_address
        wagessheet['A10'] = A10_data

        locationline = wagessheet['A11'].value
        A11_data = locationline+' '+data_wages['Unit'][0]+', '+data_wages['Branch'][0]
        wagessheet['A11'] = A11_data

        establine = wagessheet['A12'].value
        A12_data = establine+' '+data_wages['Unit'][0]+', '+data_wages['Branch'][0]
        wagessheet['A12'] = A12_data

        peline = wagessheet['A13'].value
        A13_data = peline+' '+data_wages['Unit'][0]+', '+data_wages['Branch'][0]
        wagessheet['A13'] = A13_data

        wagessheet['F4'] = 'Combined Muster Roll-cum-Register of Wages in lieu of '+month+' '+str(year)

        wagesfinalfile = os.path.join(filelocation,'Wages.xlsx')
        wagesfile.save(filename=wagesfinalfile)

    
    def create_form_H_F(form):
        if form=='FORM H':
            formHfilepath = os.path.join(karnatakafilespath,'FormH.xlsx')
        if form=='FORM F':
            formHfilepath = os.path.join(karnatakafilespath,'FormF.xlsx')
        formHfile = load_workbook(filename=formHfilepath)
        logging.info('file has sheet: '+str(formHfile.sheetnames))
        sheetformh = formHfile[form]

        
        logging.info('create columns which are now available')

        data_formH = data.copy()

        def attandance_data(employee_attendance,i):

            leavelist = list(employee_attendance.columns[(employee_attendance=='PL').iloc[i]])
            empcodeis = employee_attendance.iloc[i]['Employee Code']
            logging.info(empcodeis)
            if 'Leave Type' in leavelist:
                leavelist.remove('Leave Type')
            emp1 = pd.DataFrame(leavelist)
            
            
            if len(emp1.index)==0:
                defaultemp = {'emp':(employee_attendance).iloc[i]['Employee Code'],'startdate':0,'enddate':0,'days':0,'start_date':'-------','end_date':'-------'}
                emp1 = pd.DataFrame(defaultemp, index=[0])
                emp1.index = np.arange(1, len(emp1) + 1)
                emp1['s.no'] = emp1.index
                emp1.reset_index(drop=True, inplace=True)
                emp1['from'] = datetime.date(year,month_num,1)
                emp1['to'] = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
                emp1['totaldays'] = calendar.monthrange(year,month_num)[1]
                emp1['leavesearned'] = (employee_attendance).iloc[i]['Monthly Increment']
                emp1['leavesstart'] = float(employee_attendance.iloc[i]['Opening'])
                emp1['leavesend'] = emp1.leavesstart - emp1.days
                emp1 = emp1[["s.no","from","to","totaldays","leavesearned","leavesstart","start_date","end_date","days","leavesend"]]
            else:
                logging.info(emp1)
                emp1.columns = ['Leaves']
                emp1['emp'] = (employee_attendance).iloc[i]['Employee Code']
                emp1['Leavesdays'] = emp1.Leaves.str[5:7].astype(int)
                emp1['daysdiff'] = (emp1.Leavesdays.shift(-1) - emp1.Leavesdays).fillna(0).astype(int)
                emp1['startdate'] = np.where(emp1.daysdiff.shift() != 1, emp1.Leavesdays, 0)
                emp1['enddate'] = np.where(emp1.daysdiff!=1, emp1.Leavesdays, 0)
                emp1.drop(emp1[(emp1.startdate==0) & (emp1.enddate==0)].index, inplace=True)
                emp1['startdate'] = np.where(emp1.startdate ==0, emp1.startdate.shift(), emp1.startdate).astype(int)
                emp1['enddate'] = np.where(emp1.enddate ==0, emp1.enddate.shift(-1), emp1.enddate).astype(int)
                emp1 = emp1[['emp','startdate','enddate']]
                emp1.drop_duplicates(subset='startdate', inplace=True)
                emp1['days'] = emp1.enddate -emp1.startdate +1
                emp1['start_date'] = [datetime.date(year,month_num,x) for x in emp1.startdate]
                emp1['end_date'] = [datetime.date(year,month_num,x) for x in emp1.enddate]
                emp1.index = np.arange(1, len(emp1) + 1)
                emp1['s.no'] = emp1.index
                emp1.reset_index(drop=True, inplace=True)
                emp1['from'] = datetime.date(year,month_num,1)
                emp1['to'] = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
                emp1['totaldays'] = calendar.monthrange(year,month_num)[1]
                emp1['leavesearned'] = (employee_attendance).iloc[i]['Monthly Increment']
                emp1['totalleaves']= float(employee_attendance.iloc[i]['Opening'])
                emp1['cumdays']=emp1['days'].cumsum()
                emp1['leavesend'] = emp1.totalleaves - emp1.cumdays
                emp1['leavesstart'] =emp1['totalleaves']
                emp1 = emp1[["s.no","from","to","totaldays","leavesearned","leavesstart","start_date","end_date","days","leavesend"]]
                
            
            return emp1

        def prepare_emp_sheet(emp1,sheet_key,key,name,fathername):
            
            sheet1 = formHfile.copy_worksheet(sheetformh)
            sheet1.title = sheet_key
            lastline = sheet1['B18'].value
            sheet1['B18'] =''

            if len(emp1)>3:
                lastlinerow = 'B'+str(18+len(emp1))
            else:
                lastlinerow = 'B18'

            
            logging.info(lastlinerow)
            sheet1[lastlinerow] = lastline

            
            from openpyxl.utils.dataframe import dataframe_to_rows
            rows = dataframe_to_rows(emp1, index=False, header=False)

            for r_idx, row in enumerate(rows, 14):
                for c_idx, value in enumerate(row, 2):
                    sheet1.cell(row=r_idx, column=c_idx, value=value)
            sheet1['H5']=key
            sheet1['F7']=name
            sheet1['F8']=fathername

            sheet1.sheet_properties.pageSetUpPr.fitToPage = True

        emp_count = len(data_formH.index)
        emp_dic = dict()
        for i in range(0,emp_count):
            key = (data_formH).iloc[i]['Employee Code']
            emp_dic[key] = attandance_data(data_formH,i)
            sheet_key = form+'_'+str(key)
            name= data_formH[data_formH['Employee Code']==key]['Employee Name'].values[0]
            fathername= data_formH[data_formH['Employee Code']==key]["Father's Name"].values[0]
            logging.info(name)
            logging.info(fathername)
            prepare_emp_sheet(emp_dic[key],sheet_key,key,name,fathername)
            logging.info(key)
            logging.info(sheet_key)
        if form=='FORM H':
            formHfinalfile = os.path.join(filelocation,'FormH.xlsx')
        if form=='FORM F':
            formHfinalfile = os.path.join(filelocation,'FormF.xlsx')
        
        formHfile.remove(sheetformh)
        formHfile.save(filename=formHfinalfile)

    
    def create_muster():

        musterfilepath = os.path.join(karnatakafilespath,'Muster.xlsx')
        musterfile = load_workbook(filename=musterfilepath)
        logging.info('muster file has sheet: '+str(musterfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_muster = data.copy()

        data_muster['S.no'] = list(range(1,len(data_muster)+1))

        first3columns = ["S.no",'Employee Code','Employee Name']
        last2columns = ["Date Left","Days Paid"]

        columnstotake =[]
        days = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
        for day in days:
            for col in data_muster.columns:
                if col[5:7]==day:
                    columnstotake.append(col)
        if len(columnstotake)==28:

            columnstotake.append('29')
            columnstotake.append('30')
            columnstotake.append('31')
            data_muster['29'] = ''
            data_muster['30'] = ''
            data_muster['31'] = ''
            
        elif len(columnstotake)==29:
            columnstotake.append('30')
            columnstotake.append('31')
            data_muster['30'] = ''
            data_muster['31'] = ''

        elif len(columnstotake)==30:
            columnstotake.append('31')
            data_muster['31'] = ''

        muster_columns = first3columns+columnstotake+last2columns

        muster_data = data_muster[muster_columns]

        mustersheet = musterfile['Muster']

        mustersheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for muster is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(muster_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 18):
            for c_idx, value in enumerate(row, 2):
                mustersheet.cell(row=r_idx, column=c_idx, value=value)
                mustersheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                mustersheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                mustersheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        logging.info('')

        contractline = mustersheet['B10'].value
        B10_data = contractline+' '+contractor_name+', '+contractor_address
        mustersheet['B10'] = B10_data

        locationline = mustersheet['B11'].value
        B11_data = locationline+' '+data_muster['Unit'][0]+', '+data_muster['Branch'][0]
        mustersheet['B11'] = B11_data

        establine = mustersheet['B12'].value
        B12_data = establine+' '+data_muster['Unit'][0]+', '+data_muster['Branch'][0]
        mustersheet['B12'] = B12_data

        peline = mustersheet['B13'].value
        B13_data = peline+' '+data_muster['Unit'][0]+', '+data_muster['Branch'][0]
        mustersheet['B13'] = B13_data

        mustersheet['B4'] = 'Combined Muster Roll-cum-Register of Wages in lieu of '+month+' '+str(year)

        musterfinalfile = os.path.join(filelocation,'Muster.xlsx')
        musterfile.save(filename=musterfinalfile)

    def create_formXIX():

        formXIXfilepath = os.path.join(karnatakafilespath,'FormXIX.xlsx')
        formXIXfile = load_workbook(filename=formXIXfilepath)
        logging.info('Form XIX file has sheet: '+str(formXIXfile.sheetnames))
        sheetformXIX = formXIXfile['FORM XIX']

        
        logging.info('create columns which are now available')

        data_formXIX = data.copy()

        emp_count = len(data_formXIX.index)
        
        for i in range(0,emp_count):
            key = (data_formXIX).iloc[i]['Employee Code']
            sheet_key = 'FORM XIX_'+str(key)

            emp_data = (data_formXIX).iloc[i]
            emp_data.fillna(value='', inplace=True)

            sheet1 = formXIXfile.copy_worksheet(sheetformXIX)
            sheet1.title = sheet_key
            sheet1['D7'] = contractor_name+', '+contractor_address
            sheet1['D8'] = emp_data['Unit']+', '+emp_data['Branch']
            sheet1['D9'] = emp_data['Unit']+', '+emp_data['Branch']
            sheet1['D10'] = emp_data['Unit']+', '+emp_data['Branch']
            sheet1['D11'] = emp_data['Employee Name']
            sheet1['D12'] = emp_data['Gender']
            sheet1['D13'] = month+'-'+str(year)
            sheet1['D14'] = key
            sheet1['D15'] = emp_data['Days Paid']
            sheet1['D16'] = emp_data['Earned Basic']
            sheet1['D17'] = emp_data['HRA']
            sheet1['D18'] = emp_data['Tel and Int Reimb']
            sheet1['D19'] = emp_data['Bonus']
            sheet1['D20'] = emp_data['Fuel Reimb']
            sheet1['D21'] = emp_data['Corp Attire Reimb']
            sheet1['D22'] = emp_data['CCA']
            sheet1['D23'] = emp_data['Other Earning']
            sheet1['D24'] = emp_data['Total Earning']
            sheet1['D25'] = emp_data['Insurance']
            sheet1['D26'] = emp_data['P.Tax']
            sheet1['D27'] = emp_data['TDS']
            sheet1['D28'] = emp_data['Total Deductions']
            sheet1['D29'] = emp_data['Net Paid']

        formXIXfinalfile = os.path.join(filelocation,'FormXIX.xlsx')
        formXIXfile.remove(sheetformXIX)
        formXIXfile.save(filename=formXIXfinalfile)

    def create_ecard():

        ecardfilepath = os.path.join(karnatakafilespath,'Employment card.xlsx')
        ecardfile = load_workbook(filename=ecardfilepath)
        logging.info('Employment card file has sheet: '+str(ecardfile.sheetnames))
        sheetecard = ecardfile['Employment card']

        
        logging.info('create columns which are now available')

        data_ecard = data.copy()

        emp_count = len(data_ecard.index)
        
        for i in range(0,emp_count):
            key = (data_ecard).iloc[i]['Employee Code']
            sheet_key = 'Employment card_'+str(key)

            emp_data = (data_ecard).iloc[i]
            emp_data.fillna(value='', inplace=True)

            sheet1 = ecardfile.copy_worksheet(sheetecard)
            sheet1.title = sheet_key
            sheet1['B4'] = contractor_name
            sheet1['B5'] = ''
            sheet1['B6'] = ''
            sheet1['B7'] = ''
            sheet1['B8'] = emp_data['Department']
            sheet1['B9'] = contractor_address
            sheet1['B10'] = emp_data['Unit']
            sheet1['B11'] = emp_data['Registration_no']
            sheet1['B12'] = ''
            sheet1['B13'] = ''
            sheet1['B14'] = emp_data['Employee Name']
            sheet1['B15'] = emp_data['Aadhar Number']
            sheet1['B16'] = emp_data['Mobile Tel No.']
            sheet1['B17'] = key
            sheet1['B18'] = emp_data['Designation']
            sheet1['B19'] = emp_data['Net Paid']
            sheet1['B20'] = emp_data['Date Joined']
            

        ecardfinalfile = os.path.join(filelocation,'Employment card.xlsx')
        ecardfile.remove(sheetecard)
        ecardfile.save(filename=ecardfinalfile)
            




        
    
    create_form_A()
    create_form_B()
    create_form_XXI()
    create_form_XXII()
    create_form_XXIII()
    create_form_XX()
    create_wages()
    create_form_H_F('FORM H')
    create_form_H_F('FORM F')
    create_muster()
    create_formXIX()
    create_ecard()





Stateslist = ['Karnataka','Maharashtra','Delhi','Telangana','Uttar Pradesh','Tamilnadu']

State_Process = {'Karnataka':Karnataka,'Maharashtra':Maharashtra,'Delhi':Delhi,'Telangana':Telangana,'Uttar Pradesh':Uttar_Pradesh,'Tamilnadu':Tamilnadu}

companylist = ['SVR LTD','PRY Wine Ltd','CDE Technology Ltd']



def svrDataProcess(inputfolder,month,year):
    logging.info('scr data process running')

def pryDataProcess(inputfolder,month,year):
    logging.info('pry data process running')

    emp_df_columns = ['Employee Code', 'Employee Name', 'Company Name', 'Grade', 'Branch',
       'Department', 'Designation', 'Division', 'Group', 'Category', 'Unit',
       'Location Code', 'State', 'Date of Birth', 'Date Joined',
       'Date of Confirmation', 'Date Left', 'Title', 'Last Inc. Date',
       'Ticket Number', 'Local Address 1', 'Local Address 2',
       'Local Address 3', 'Local Address 4', 'Local City Name',
       'Local District Name', 'Local PinCode', 'Local State Name',
       'Residence Tel No.', 'Permanent Address 1', 'Permanent Address 2',
       'Permanent Address 3', 'Permanent Address 4', 'UAN Number',
       'Permanent Tel No.', 'Office Tel No.', 'Extension Tel No.',
       'Mobile Tel No.', "Father's Name", 'Gender', 'Age', 'Number of Months',
       'Marital Status', 'PT Number', 'PF Number (Old Version)', 'PF Number',
       'PF Number (WithComPrefix)', 'PAN Number', 'ESIC Number (Old Version)',
       'ESIC Number', 'ESIC Number (CompPrefix)', 'FPF Number', 'PF Flag',
       'ESIC Flag', 'PT Flag', 'Bank A/c Number', 'Bank Name', 'Mode',
       'Account Code', 'E-Mail', 'Remarks', 'PF Remarks', 'ESIC Remarks',
       'ESIC IMP Code', 'ESIC IMP Name', 'Employee Type (For PF)',
       'Freeze Account', 'Freeze Date', 'Freeze Reason', 'Type of House (In)',
       'Comp. adn.', 'Staying In (Metro Type)', 'Children (For CED)',
       'TDS Rate', 'Resignation Date', 'Reason for Leaving', 'Bank A/C No.1',
       'Bank A/C No.2', 'Bank A/C No.3', 'Alt.Email', 'Emp Status',
       'Probation Date', 'Surcharge Flag', 'Gratuity Code',
       'Resign Offer Date', 'Permanent City', 'Permanent District',
       'Permanent Pin Code', 'Permanent State', 'Spouse Name',
       'PF Joining Date', 'PRAN Number', 'Group Joining Date', 'Aadhar Number',
       'Child in Hostel (For CED)', 'Total Exp in Years', 'P', 'L']

    sal_df_columns = ['PAY_DATE', 'EMPCODE', 'EMPNAME', 'DOJ', 'GENDER', 'DOB', 'DOL',
       'OT HOURS', 'STDDAYS', 'LOP DAYS', 'WRKDAYS', 'BASIC+DA',
       'BASIC+DA ARREAR', 'HOUSE RENT ALLOWANCE', 'HRA ARREAR',
       'MEDICAL ALLOWANCE', 'MEDICAL ALLOWANCE ARREARS',
       'CHILD EDUCATION ALLOWANCE', 'CHILD EDUCATION ALL ARREARS',
       'HELPER ALLOWANCE', 'HELPER ALLOWANCE ARREARS', 'OTHER ALLOWANCE',
       'OTHER ALLOWANCE  ARREARS', 'OTHER EARNINGS NON TAXABLE',
       'OTHER EARNINGS', 'SALES COMMISSION', 'PROFITABILITY & SALES',
       'MBO PAYMENT', 'STAT BONUS PAYMENT', 'VEHICLE SUBSIDY', 'EXGRATIA',
       'NOTICE PAY BUYOUTS', 'OVERTIME', 'MARRIAGE GIFT', 'REFERRAL BONUS',
       'FESTIVAL ADVANCE PAY', 'LONG TERM SERVICE AWARD', 'LEAVE ENCASHMENT',
       'COMPANY SALES PAYMENT', 'PROFITABILITY PAYMENT',
       'Ticket Restaurant Meal Card', 'UNIFORM CLEARANCE', 'LTA REIMBURSEMENT',
       'GRAPE QUALITY PAY', 'SHORT PAY EARNING',
       'HOSPITALITY PERFORMANCE BONUS', 'OVERSEAS SPECIAL ALLOWANCE',
       'GROSS_EARN', 'Professional Tax', 'ESI']

    file_list = os.listdir(inputfolder)
    logging.info('input folder is '+str(inputfolder))
    for f in file_list:
        if f[0:6].upper()=='MASTER':
            masterfilename = f
            logging.info('masterfilename is :'+f)
        if f[0:6].upper()=='SALARY':
            salaryfilename = f
            logging.info('salaryfilename is :'+f)
        if f[0:10].upper()=='ATTENDANCE':
            attendancefilename = f
            logging.info('attendancefilename is :'+f)
        if f[0:5].upper()=='LEAVE':
            leavefilename = f
            logging.info('leavefilename is :'+f)
        if f[0:14].upper()=='LEFT EMPLOYEES':
            leftempfilename = f
            logging.info('leftempfilename is :'+f)
        if f[0:9].upper()=='CDE UNITS':
            unitfilename = f
            logging.info('unitfilename is :'+f)
    
    logging.info('file names set')
    
    if 'masterfilename' in locals():
        masterfile = os.path.join(inputfolder,masterfilename)
        employee_data = pd.read_excel(masterfile)
        employee_data.dropna(how='all', inplace=True)
        employee_data.reset_index(drop=True, inplace=True)
        employee_data.columns = employee_data.iloc[0]
        employee_data.drop(0, inplace=True)
        logging.info('employee data loaded')
    else:
        employee_data = pd.DataFrame(columns = emp_df_columns)
        logging.error('employee data not available setting empty dataset')
    if 'salaryfilename' in locals():
        salaryfile = os.path.join(inputfolder,salaryfilename)
        salary_data = pd.read_excel(salaryfile)
    else:
        salary_data = pd.DataFrame(columns= sal_df_columns)
        logging.error('salary data not available setting empty dataset')
    if 'attendancefilename' in locals():
        attendancefile = os.path.join(inputfolder,attendancefilename)
        attendance_data_1 = pd.read_excel(attendancefile, sheet='HO')
        attendance_data_2 = pd.read_excel(attendancefile, sheet='ROI')
        attendance_data = attendance_data_1.concat(attendance_data_2)
    else:
        attendance_data = pd.DataFrame(columns= sal_df_columns)
        logging.error('attendance data not available setting empty dataset')

    
    pry_data = salary_data.merge(attendance_data, how='left', left_on='EMPCODE', right_on='Empl Code')




def cdeDataProcess(inputfolder,month,year):
    global nomatch
    nomatch=''
    logging.info('cde data process running')

    emp_df_columns = ['Employee Code', 'Employee Name', 'Company Name', 'Grade', 'Branch',
       'Department', 'Designation', 'Division', 'Group', 'Category', 'Unit',
       'Location Code', 'State', 'Date of Birth', 'Date Joined',
       'Date of Confirmation', 'Date Left', 'Title', 'Last Inc. Date',
       'Ticket Number', 'Local Address 1', 'Local Address 2',
       'Local Address 3', 'Local Address 4', 'Local City Name',
       'Local District Name', 'Local PinCode', 'Local State Name',
       'Residence Tel No.', 'Permanent Address 1', 'Permanent Address 2',
       'Permanent Address 3', 'Permanent Address 4', 'UAN Number',
       'Permanent Tel No.', 'Office Tel No.', 'Extension Tel No.',
       'Mobile Tel No.', "Father's Name", 'Gender', 'Age', 'Number of Months',
       'Marital Status', 'PT Number', 'PF Number (Old Version)', 'PF Number',
       'PF Number (WithComPrefix)', 'PAN Number', 'ESIC Number (Old Version)',
       'ESIC Number', 'ESIC Number (CompPrefix)', 'FPF Number', 'PF Flag',
       'ESIC Flag', 'PT Flag', 'Bank A/c Number', 'Bank Name', 'Mode',
       'Account Code', 'E-Mail', 'Remarks', 'PF Remarks', 'ESIC Remarks',
       'ESIC IMP Code', 'ESIC IMP Name', 'Employee Type (For PF)',
       'Freeze Account', 'Freeze Date', 'Freeze Reason', 'Type of House (In)',
       'Comp. adn.', 'Staying In (Metro Type)', 'Children (For CED)',
       'TDS Rate', 'Resignation Date', 'Reason for Leaving', 'Bank A/C No.1',
       'Bank A/C No.2', 'Bank A/C No.3', 'Alt.Email', 'Emp Status',
       'Probation Date', 'Surcharge Flag', 'Gratuity Code',
       'Resign Offer Date', 'Permanent City', 'Permanent District',
       'Permanent Pin Code', 'Permanent State', 'Spouse Name',
       'PF Joining Date', 'PRAN Number', 'Group Joining Date', 'Aadhar Number',
       'Child in Hostel (For CED)', 'Total Exp in Years', 'P', 'L']

    salary_df_columns = ['Sr', 'DivisionName', 'Sal Status', 'Emp Code', 'Emp Name', 'DesigName',
       'Date Joined', 'UnitName', 'Branch', 'Days Paid', 'Earned Basic', 'HRA',
       'Conveyance', 'Medical Allowance', 'Telephone Reimb',
       'Tel and Int Reimb', 'Bonus', 'Other Allowance', 'Fuel Reimb',
       'Prof Dev Reimb', 'Corp Attire Reimb', 'Meal Allowance',
       'Special Allowance', 'Personal Allowance', 'CCA', 'Other Reimb',
       'Arrears', 'Other Earning', 'Variable Pay', 'Leave Encashment',
       'Stipend', 'Consultancy Fees', 'Total Earning', 'Insurance', 'CSR',
       'PF', 'ESIC', 'P.Tax', 'LWF EE', 'Salary Advance', 'Loan Deduction',
       'Loan Interest', 'Other Deduction', 'TDS', 'Total Deductions',
       'Net Paid', 'BankName', 'Bank A/c Number', 'Account Code', 'Remarks',
       'PF Number (Old)', 'UAN Number', 'ESIC Number', 'Personal A/c Number',
       'E-Mail', 'Mobile No.', 'FIXED MONTHLY GROSS', 'CHECK CTC Gross']

    atten_df_columns = ['Emp Code', 'Employee Name', 'Branch', 'Designation', 'Sat\r\n01/02',
       'Sun\r\n02/02', 'Mon\r\n03/02', 'Tue\r\n04/02', 'Wed\r\n05/02',
       'Thu\r\n06/02', 'Fri\r\n07/02', 'Sat\r\n08/02', 'Sun\r\n09/02',
       'Mon\r\n10/02', 'Tue\r\n11/02', 'Wed\r\n12/02', 'Thu\r\n13/02',
       'Fri\r\n14/02', 'Sat\r\n15/02', 'Sun\r\n16/02', 'Mon\r\n17/02',
       'Tue\r\n18/02', 'Wed\r\n19/02', 'Thu\r\n20/02', 'Fri\r\n21/02',
       'Sat\r\n22/02', 'Sun\r\n23/02', 'Mon\r\n24/02', 'Tue\r\n25/02',
       'Wed\r\n26/02', 'Thu\r\n27/02', 'Fri\r\n28/02', 'Sat\r\n29/02',
       'Total\r\nDP', 'Total\r\nABS', 'Total\r\nLWP', 'Total\r\nCL',
       'Total\r\nSL', 'Total\r\nPL', 'Total\r\nL1', 'Total\r\nL2',
       'Total\r\nL3', 'Total\r\nL4', 'Total\r\nL5', 'Total\r\nCO-',
       'Total\r\nCO+', 'Total\r\nOL', 'Total\r\nWO', 'Total\r\nPH',
       'Total\r\nEO', 'Total\r\nWOP', 'Total\r\nPHP', 'Total\r\nOT Hrs',
       'Total\r\nLT Hrs']

    leave_df_columns = ['Emp. Code', 'Emp. Name', 'Leave Type', 'Opening', 'Monthly Increment',
       'Used', 'Closing', 'Leave Accrued', 'Encash']

    leftemp_df_columns = ['Employee Name', 'Employee Code', 'Date Joined', 'Date Left',
       'UAN Number']

    unit_df_columns = ['Unit', 'Location_code','Location', 'Address', 'Registration_no', 'PE_or_contract',
       'State_or_Central', 'start_time', 'end_time', 'rest_interval','Contractor_name','Contractor_Address']

    logging.info('column variables set')

    

    
    file_list = os.listdir(inputfolder)
    logging.info('input folder is '+str(inputfolder))
    for f in file_list:
        if f[0:6].upper()=='MASTER':
            masterfilename = f
            logging.info('masterfilename is :'+f)
        if f[0:6].upper()=='SALARY':
            salaryfilename = f
            logging.info('salaryfilename is :'+f)
        if f[0:10].upper()=='ATTENDANCE':
            attendancefilename = f
            logging.info('attendancefilename is :'+f)
        if f[0:5].upper()=='LEAVE':
            leavefilename = f
            logging.info('leavefilename is :'+f)
        if f[0:14].upper()=='LEFT EMPLOYEES':
            leftempfilename = f
            logging.info('leftempfilename is :'+f)
        if f[0:9].upper()=='CDE UNITS':
            unitfilename = f
            logging.info('unitfilename is :'+f)
    
    logging.info('file names set')
    
    if 'masterfilename' in locals():
        masterfile = os.path.join(inputfolder,masterfilename)
        employee_data = pd.read_excel(masterfile)
        employee_data.dropna(how='all', inplace=True)
        employee_data.reset_index(drop=True, inplace=True)
        logging.info('employee data loaded')
    else:
        employee_data = pd.DataFrame(columns = emp_df_columns)
        logging.error('employee data not available setting empty dataset')
    if 'salaryfilename' in locals():
        salaryfile = os.path.join(inputfolder,salaryfilename)
        salary_data = pd.read_excel(salaryfile)
        salary_data.dropna(how='all', inplace=True)
        salary_data.reset_index(drop=True, inplace=True)
        logging.info('salary data loaded')
    else:
        salary_data = pd.DataFrame(columns = salary_df_columns)
        logging.info('salary data not available setting empty dataset')
    if 'attendancefilename' in locals():
        attendancefile = os.path.join(inputfolder,attendancefilename)
        attendance_data = pd.read_excel(attendancefile)
        attendance_data.dropna(how='all', inplace=True)
        attendance_data.reset_index(drop=True, inplace=True)
        logging.info('attendance data loaded')
    else:
        attendance_data = pd.DataFrame(columns = atten_df_columns)
        logging.info('attendance data not available setting empty dataset')
    if 'leavefilename' in locals():
        leavefile = os.path.join(inputfolder,leavefilename)
        leave_data = pd.read_excel(leavefile)
        leave_data.dropna(how='all', inplace=True)
        leave_data.reset_index(drop=True, inplace=True)
        logging.info('leave data loaded')
    else:
        leave_data = pd.DataFrame(columns = leave_df_columns)
        logging.info('leave data not available setting empty dataset')
    if 'leftempfilename' in locals():
        leftempfile = os.path.join(inputfolder,leftempfilename)
        leftemp_data = pd.read_excel(leftempfile)
        leftemp_data.dropna(how='all', inplace=True)
        leftemp_data.reset_index(drop=True, inplace=True)
        logging.info('left employees data loaded')
    else:
        leftemp_data = pd.DataFrame(columns = leftemp_df_columns)
        logging.info('left employees data not available setting empty dataset')
    if 'unitfilename' in locals():
        unitfile = os.path.join(inputfolder,unitfilename)
        unit_data = pd.read_excel(unitfile)
        unit_data.dropna(how='all', inplace=True)
        unit_data.reset_index(drop=True, inplace=True)
        logging.info('unit data loaded')
    else:
        unit_data = pd.DataFrame(columns = unit_df_columns)
        logging.info('unit data not available setting empty dataset')

    employee_data.drop(columns='Date Left', inplace=True)

    employee_data['Location Code'] = employee_data['Location Code'].astype(int)
    employee_data['Employee Code'] = employee_data['Employee Code'].astype(str)

    unit_data['Location Code'] = unit_data['Location Code'].astype(int)

    salary_data.drop(columns=list(employee_data.columns.intersection(salary_data.columns)), inplace=True)

    salary_data['Emp Code'] = salary_data['Emp Code'].astype(str)

    attendance_data.drop(columns=['Employee Name', 'Branch', 'Designation'], inplace=True)
    attendance_data['Emp Code'] = attendance_data['Emp Code'].astype(str)

    leave_data['Emp. Code'] = leave_data['Emp. Code'].astype(str)

    leftemp_data.drop(columns=['Employee Name', 'Date Joined', 'UAN Number'],inplace=True)

    leftemp_data['Employee Code'] = leftemp_data['Employee Code'].astype(str)


    
    
    CDE_Data = employee_data.merge(unit_data,how='left',on='Location Code').merge(
        salary_data,how='left',left_on='Employee Code',right_on='Emp Code').merge(
            attendance_data,how='left',left_on='Employee Code', right_on='Emp Code').merge(
                leave_data, how='left', left_on='Employee Code', right_on='Emp. Code').merge(
                    leftemp_data, how='left', on='Employee Code')
    
    logging.info('merged all data sets')

    rename_list=[]
    renamed=[]
    drop_list=[]
    for x in list(CDE_Data.columns):
        if x[-2:]=='_x':
            rename_list.append(x)
            renamed.append(x[0:-2])
        if x[-2:]=='_y':
            drop_list.append(x)
    
    rename_dict = dict(zip(rename_list,renamed))

    CDE_Data.rename(columns=rename_dict, inplace=True)

    logging.info('columns renamed correctly')

    CDE_Data.drop(columns=drop_list, inplace=True)

    logging.info('dropped duplicate columns')

    monthyear = month+' '+str(year)
    if monthyear.upper() in masterfilename.upper():
        logging.info('month year matches with data')

        #for all state employees(PE+contractor)

        statedata = CDE_Data[CDE_Data['State_or_Central']=='State'].copy()

        CDE_States = list(statedata['State'].unique())

        for state in CDE_States:
            

            unit_with_location = list((statedata[statedata.State==state]['Unit']+','+statedata[statedata.State==state]['Location']).unique())
            for UL in unit_with_location:
                UL=re.sub("(\.*\s+)$","",UL)
                inputdata = statedata[(statedata['State']==state) & (statedata['Unit']==UL.split(',')[0]) & (statedata['Location']==UL.split(',')[1])].copy()
                inputdata['Contractor_name'] = inputdata['Contractor_name'].fillna(value='')
                inputdata['Contractor_Address'] = inputdata['Contractor_Address'].fillna(value='')
                inpath = os.path.join(inputfolder,'Registers','States',state,UL)
                if os.path.exists(inpath):
                    logging.info('running state process')
                    contractor_name= inputdata['Contractor_name'].unique()[0]
                    contractor_address= inputdata['Contractor_Address'].unique()[0]
                    State_Process[state](data=inputdata,contractor_name=contractor_name,contractor_address=contractor_address,filelocation=inpath,month=month,year=year)
                else:
                    logging.info('making directory')
                    os.makedirs(inpath)
                    logging.info('directory created')
                    contractor_name= inputdata['Contractor_name'].unique()[0]
                    contractor_address= inputdata['Contractor_Address'].unique()[0]
                    State_Process[state](data=inputdata,contractor_name=contractor_name,contractor_address=contractor_address,filelocation=inpath,month=month,year=year)

        #for contractors form
        contractdata = CDE_Data[(CDE_Data['State_or_Central']=='State') & (CDE_Data['PE_or_contract']=='Contract')].copy()
        contractor_units = list((contractdata['Unit']+','+contractdata['Location']).unique())
        for UL in contractor_units:
            inputdata = contractdata[(contractdata['Unit']==UL.split(',')[0]) & (contractdata['Location']==UL.split(',')[1])]
            
            inpath = os.path.join(inputfolder,'Registers','Contractors',UL)
            if os.path.exists(inpath):
                logging.info('running contractor process')
            else:
                logging.info('making directory')
                os.makedirs(inpath)
                logging.info('directory created')

        #for central form
        centraldata = CDE_Data[CDE_Data['State_or_Central']=='Central'].copy()
        central_units = list((centraldata['Unit']+','+centraldata['Location']).unique())
        for UL in central_units:
            inputdata = centraldata[(centraldata['Unit']==UL.split(',')[0]) & (centraldata['Location']==UL.split(',')[1])]
            
            inpath = os.path.join(inputfolder,'Registers','Central',UL)
            if os.path.exists(inpath):
                logging.info('running contractor process')
            else:
                logging.info('making directory')
                os.makedirs(inpath)
                logging.info('directory created')
    
    else:
        nomatch = "Date you mentioned doesn't match with Input data"
        logging.error(nomatch)

    

DataProcess = {'SVR LTD':svrDataProcess,'PRY Wine Ltd':pryDataProcess,'CDE Technology Ltd':cdeDataProcess}


def CompanyDataProcessing(company,inputfolder,month,year):
    inputfolder = Path(inputfolder)
    yr = int(year)
    DataProcess[company](inputfolder,month,yr)

#backend code ends here

companies = ['SVR LTD','PRY Wine Ltd','CDE Technology Ltd']

Months = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']

Years = ['2017','2018','2019','2020']

companyname = tk.StringVar()

month = tk.StringVar()
year = tk.StringVar()


folderLabel = ttk.LabelFrame(master, text="Select the Company")
folderLabel.grid(column=0,row=1,padx=20,pady=20)

companynameLabel = Label(folderLabel, text="Company Name")
companynameLabel.grid(column=1, row=0, padx=20,pady=20)

comapnynameEntry = ttk.Combobox(folderLabel,values=companies,textvariable=companyname)
comapnynameEntry.grid(column=2, row=0, padx=20,pady=20)

MonthLabel = Label(folderLabel, text="Month and Year")
MonthLabel.grid(column=1, row=2, padx=20,pady=20)

MonthEntry = ttk.Combobox(folderLabel,values=Months,textvariable=month)
MonthEntry.grid(column=2, row=2, padx=20,pady=20)

YearEntry = ttk.Combobox(folderLabel,values=Years,textvariable=year)
YearEntry.grid(column=3, row=2, padx=20,pady=20)

def disfo():
    foldername = filedialog.askdirectory()
    logging.info(foldername)
    logging.info(type(foldername))
    foldernamelabel.configure(text=foldername)


button = ttk.Button(folderLabel, text = "Select Company Folder", command=disfo)
button.grid(column=1, row=1, columnspan=2,padx=20, pady=20)

foldernamelabel = Label(folderLabel, text="")
foldernamelabel.grid(column=1, row=3, columnspan=2,padx=20,pady=20)




def generateforms(comp,mn,yr):
    company=comp.get()

    month = mn.get()
    year = yr.get()


    getfolder = foldernamelabel.cget("text")



    logging.info(type(company))
    logging.info(company)

    logging.info(type(getfolder))
    logging.info(getfolder)

    if (company =="" and getfolder =="" and (month =="" or year =="")):
        report.configure(text="Please select month year, company folder and company name")
    elif (company=="" and getfolder =="" and not(month =="" or year =="")):
        report.configure(text="Please select company folder and company name")
    elif (company=="" and getfolder !="" and not(month =="" or year =="")):
        report.configure(text="Please select company name")
    elif (company!="" and getfolder =="" and not(month =="" or year =="")):
        report.configure(text="Please select company folder")
    elif (company =="" and getfolder !="" and (month =="" or year=="")):
        report.configure(text="Please select month year and company name")
    elif (company!="" and getfolder=="" and (month =="" or year=="")):
        report.configure(text="Please select month year and company folder")
    elif (company!="" and getfolder!="" and (month =="" or year=="")):
        report.configure(text="Please select month year")
    else:
        logging.info(company, getfolder,  month,  year)
        report.configure(text="Processing")
        try:
            CompanyDataProcessing(company,getfolder,month,year)
        except Exception as e:
            logging.info('Failed')
            report.configure('Failed')
        else:
            if nomatch=='':
                logging.info('Completed Form Creation')
                report.configure(text='Completed Form Creation')
            else:
                logging.info(nomatch)
                report.configure(text=nomatch)
        finally:
            logging.info('done')
        
def convert_forms_to_pdf():

    getfolder = foldernamelabel.cget("text")

    if getfolder=="":
        report.configure(text="Please select company folder")
    else:
        registerfolder = os.path.join(Path(getfolder),'Registers')
        if os.path.exists(registerfolder):
            for root, dirs, files in os.walk(registerfolder):
                for fileis in files:
                    if fileis.endswith(".xlsx"):
                        try:
                            create_pdf(root,fileis)
                        except Exception as e:
                            logging.info('Failed pdf Conversion')
                            report.configure(text="Failed")
                        else:
                            logging.info('Completed pdf Conversion')
                            report.configure(text="Completed")
                        finally:
                            logging.info('done')
        else:
            report.configure(text="Registers not available")
                        



generateforms = partial(generateforms,companyname,month,year)

button = ttk.Button(master, text = "Generate Forms", command=generateforms)
button.grid(column=1, row=1, columnspan=2,padx=20, pady=20)

Detailbox = ttk.LabelFrame(master, text="")
Detailbox.grid(column=0,row=2,padx=20,pady=20)

report = Label(Detailbox, text="                                                            ")
report.grid(column=0, row=0, padx=20,pady=20)



button2 = ttk.Button(master, text = "Convert forms to PDF", command=convert_forms_to_pdf)
button2.grid(column=0, row=3, columnspan=2,padx=20, pady=20)


mainloop()


