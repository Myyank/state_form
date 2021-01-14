from states import logging,monthdict,Statefolder
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging


def Goa(data,contractor_name,contractor_address,filelocation,month,year):
    Goafilespath = os.path.join(Statefolder,'Goa')
    logging.info('Goa files path is :'+str(Goafilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]

    def Form_I():

        formIfilepath = os.path.join(Goafilespath,'Form I register of Fine.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formI = data.copy()
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","name&date_of_offence","cause_against_fine","FIXED MONTHLY GROSS",
                                        "Date of payment ","Date of payment ","remarks"]

        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        data_formI[["name&date_of_offence","cause_against_fine","remarks"]]="-----"
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        formIsheet['A4']=formIsheet['A4'].value+" : "+data_formI['Unit'][0]
        formIfinalfile = os.path.join(filelocation,'Form I register of Fine.xlsx')
        formIfile.save(filename=formIfinalfile)

    def Form_II():
        formIIfilepath = os.path.join(Goafilespath,'Form II register of damage or loss.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy()
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","attendancefile",
                                        "damage_or_loss","whether_work_showed_cause",
                                        "Date of payment ","num_instalments","Date of payment ","remarks"]

        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII["attendancefile"]="confusing mapping"
        data_formII[["damage_or_loss","whether_work_showed_cause","num_instalments","remarks"]]="-----"
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formII_data, index=False, header=False)

        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        formIIsheet['A4']=formIIsheet['A4'].value+" : "+data_formII['Unit'][0]
        formIIfinalfile = os.path.join(filelocation,'Form II register of damage or loss.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_VIII():
        formVIIIfilepath = os.path.join(Goafilespath,'Form VIII register of Over time.xlsx')
        formVIIIfile = load_workbook(filename=formVIIIfilepath)
        logging.info('Form VIII file has sheet: '+str(formVIIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formVIII = data.copy()
        data_formVIII['Designation_Dept']=data_formVIII["Designation"]+"_"+data_formVIII["Department"]
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation_Dept","attendancefile",
                                        "extent_of_overtime","total_overtime",
                                        'Normal hrs ','FIXED MONTHLY GROSS',
                                        "overtime_rate",'Overtime',"ot",'CHECK CTC Gross','Date of payment ']

        data_formVIII['S.no'] = list(range(1,len(data_formVIII)+1))
        data_formVIII[["overtime_rate","attendancefile","overtime_rate","ot"]]="Didn't find"
        data_formVIII[["extent_of_overtime","total_overtime"]]="----"
        
        formVIII_data=data_formVIII[columns]
        formVIIIsheet = formVIIIfile['Sheet1']
        formVIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form VIII is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formVIII_data, index=False, header=False)

        logging.info('rows taken out from data')
        
        row_copy=dataframe_to_rows(formVIII_data, index=False, header=False)
        for i in range(len(list(row_copy))-2):
            i+=12
            formVIIIsheet.merge_cells('C'+str(i)+':D'+str(i))
            formVIIIsheet.merge_cells('F'+str(i)+':H'+str(i))
            formVIIIsheet.merge_cells('I'+str(i)+':K'+str(i))
            formVIIIsheet.merge_cells('L'+str(i)+':N'+str(i))
            formVIIIsheet.merge_cells('O'+str(i)+':R'+str(i))
            formVIIIsheet.merge_cells('S'+str(i)+':T'+str(i))
            formVIIIsheet.merge_cells('U'+str(i)+':V'+str(i))
            formVIIIsheet.merge_cells('W'+str(i)+':X'+str(i))
            formVIIIsheet.merge_cells('Y'+str(i)+':Z'+str(i))
            formVIIIsheet.merge_cells('AA'+str(i)+':AB'+str(i))
            formVIIIsheet.merge_cells('AC'+str(i)+':AD'+str(i))
            formVIIIsheet.merge_cells('AE'+str(i)+':AG'+str(i))
        
        c_idx=0
        for r_idx, row in enumerate(rows, 10):
            row_iterator=zip(row)
            while True:
                c_idx+=1
                if type(formVIIIsheet.cell(row=r_idx, column=c_idx)).__name__ == 'MergedCell':
                    continue
                try:
                    value=next(row_iterator)[0]
                except:
                    c_idx=0
                    break
                formVIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formVIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formVIIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                formVIIIfinalfile = os.path.join(filelocation,'Form VIII register of Over time.xlsx')
                formVIIIfile.save(filename=formVIIIfinalfile)
        
        formVIIIsheet['Q4']="Month ending "+month+" "+str(year)
        formVIIIfinalfile = os.path.join(filelocation,'Form VIII register of Over time.xlsx')
        formVIIIfile.save(filename=formVIIIfinalfile)
        
    #Not complete since not all mapping is provided
    def From_XII():
        formXIIfilepath = os.path.join(Goafilespath,'Form XII Register of leave.xlsx')
        formXIIfile = load_workbook(filename=formXIIfilepath)
        logging.info('Form XII file has sheet: '+str(formXIIfile.sheetnames))
        #print(formXIIfile.sheetnames)
        logging.info('create columns which are now available')

        data_formXII = data.copy()
        #print(sorted(list(data_formXII.columns)))

        


    def Form_XXI():
        formXXIfilepath = os.path.join(Goafilespath,'Form XXI Register of Employment.xlsx')
        formXXIfile = load_workbook(filename=formXXIfilepath)
        logging.info('Form XXI file has sheet: '+str(formXXIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formXXI = data.copy()
        
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation","Date_of_appoinment"]
        
        data_formXXI_columns=list(data_formXXI.columns)
        start=data_formXXI_columns.index('Arrears salary')
        end=data_formXXI_columns.index('Total\r\nDP')
        columns.extend(data_formXXI_columns[start+1:end])
        
        less=31-len(data_formXXI_columns[start+1:end])
        for i in range(less):
            columns.extend(["less"+str(i+1)])
            data_formXXI["less"+str(i+1)]=""

        columns.extend(["normal_hours",'Overtime',"remarks"])
        data_formXXI[["normal_hours","remarks","Date_of_appoinment"]]="didn't get mapping"
        data_formXXI['S.no'] = list(range(1,len(data_formXXI)+1))

        formXXI_data=data_formXXI[columns]
        formXXIsheet = formXXIfile['Sheet1']
        formXXIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form XXI is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXXI_data, index=False, header=False)

        logging.info('rows taken out from data')
        formXXIsheet.unmerge_cells('A23:E23')
        for r_idx, row in enumerate(rows, 14):
            for c_idx, value in enumerate(row, 1):
                formXXIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        formXXIsheet['AE4']=formXXIsheet['AE4'].value+"   "+str(data_formXXI['Registration_no'].unique()[0])
        formXXIsheet['A4']=formXXIsheet['A4'].value+" "+str(data_formXXI['Unit'].unique()[0])+", "+str(data_formXXI['Location'].unique()[0])
        formXXIsheet['A5']=formXXIsheet['A5'].value+" "+str(data_formXXI['Unit'].unique()[0])+", "+str(data_formXXI['Location'].unique()[0])
        formXXIfinalfile = os.path.join(filelocation,'Form XXI register of Over time.xlsx')
        formXXIfile.save(filename=formXXIfinalfile)



    def Form_XXIII():
        formXXIIIfilepath = os.path.join(Goafilespath,'Form XXIII Register of wages.xlsx')
        formXXIIIfile = load_workbook(filename=formXXIIIfilepath)
        logging.info('Form XXIII file has sheet: '+str(formXXIIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formXXIII = data.copy()
        
        columns=['S.no',"Employee Name","Father's Name","Designation",'Basic','Dearness_Allowance',
                                'Earned Basic','Dearness_Allowance_2','Other Allowance','Overtime',
                                 'FIXED MONTHLY GROSS','Salary Advance','PF', 'Other Deduction',
                                 'Total Deductions','Net Paid',"sign",'Date of payment ']
        
        data_formXXIII["sign"]=""
        data_formXXIII[["Basic",'Dearness_Allowance','Dearness_Allowance_2',"remarks","Date_of_appoinment"]]="didn't get mapping"
        data_formXXIII['S.no'] = list(range(1,len(data_formXXIII)+1))

        formXXIII_data=data_formXXIII[columns]
        formXXIIIsheet = formXXIIIfile['Sheet1']
        formXXIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form XXIII is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXXIII_data, index=False, header=False)

        logging.info('rows taken out from data')
        formXXIIIsheet.unmerge_cells('P15:R15')
        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                #formXXIIIsheet.cell(row=r_idx, column=c_idx).value=value
                formXXIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        
        formXXIIIsheet['P'+str(len(list(rows))+10+5)].value="Signature of Employer"
        
        formXXIIIsheet.merge_cells('P'+str(len(list(rows))+10+5)+':R'+str(len(list(rows))+10+5))
        
        formXXIIIsheet['P4']=formXXIIIsheet['P4'].value+"   "+str(data_formXXIII['Registration_no'].unique()[0])
        formXXIIIsheet['P5']=formXXIIIsheet['P5'].value+"   "+month
        formXXIIIsheet['A4']=formXXIIIsheet['A4'].value+" "+str(data_formXXIII['Unit'].unique()[0])
        formXXIIIsheet['A5']=formXXIIIsheet['A5'].value+" "+str(data_formXXIII['Unit'].unique()[0])+", "+str(data_formXXIII['Address'].unique()[0])
        
        formXXIIIfinalfile = os.path.join(filelocation,'Form XXIII Register of wages.xlsx')
        formXXIIIfile.save(filename=formXXIIIfinalfile)
        
    Form_I()
    Form_II()
    Form_VIII()
    From_XII()
    Form_XXI()
    Form_XXIII()